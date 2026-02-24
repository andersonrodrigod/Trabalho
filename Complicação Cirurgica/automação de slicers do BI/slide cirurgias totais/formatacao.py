from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


AZUL = "2652B5"
BRANCO = "FFFFFF"
CINZA = "EDEDED"
PRETO = "000000"


def _formatar_aba(ws) -> None:

    # Titulos finais
    ws.cell(row=1, column=1, value="ESPECIALISTA")
    ws.cell(row=1, column=2, value="TOTAL DE CIRURGIAS REALIZADAS")
    ws.cell(row=1, column=3, value='Nº BENEF "SIM"')
    ws.cell(row=1, column=4, value="PROPORCIONALIDADE")
    ws.cell(row=1, column=5, value="REPRESENTATIVIDADE")

    max_row = ws.max_row
    max_col = 5

    fill_azul = PatternFill(fill_type="solid", fgColor=AZUL)
    fill_branco = PatternFill(fill_type="solid", fgColor=BRANCO)
    fill_cinza = PatternFill(fill_type="solid", fgColor=CINZA)

    fonte_branca = Font(color=BRANCO, bold=True)
    fonte_preta_negrito = Font(color=PRETO, bold=True)

    alinhamento_centro = Alignment(horizontal="center", vertical="center")
    alinhamento_esquerda = Alignment(horizontal="left", vertical="center")

    borda_branca = Border(
        left=Side(style="thin", color=BRANCO),
        right=Side(style="thin", color=BRANCO),
        top=Side(style="thin", color=BRANCO),
        bottom=Side(style="thin", color=BRANCO),
    )

    borda_padrao = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    # Cabecalho azul com fonte branca
    for col in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = fill_azul
        cell.font = fonte_branca
        cell.alignment = alinhamento_centro
        cell.border = borda_branca

    # Descobre ultima linha TOTAL (esperada no final)
    linha_total = None
    for row in range(2, max_row + 1):
        if str(ws.cell(row=row, column=1).value).strip().upper() == "TOTAL":
            linha_total = row

    # Formata linhas de dados
    for row in range(2, max_row + 1):
        eh_total = linha_total is not None and row == linha_total

        # Coluna ESPECIALISTA sempre azul
        cell_especialista = ws.cell(row=row, column=1)
        cell_especialista.fill = fill_azul
        cell_especialista.font = fonte_branca
        cell_especialista.alignment = alinhamento_esquerda
        cell_especialista.border = borda_branca

        # Linha TOTAL inteira azul, branca e negrito
        if eh_total:
            for col in range(2, max_col + 1):
                c = ws.cell(row=row, column=col)
                c.fill = fill_azul
                c.font = fonte_branca
                c.alignment = alinhamento_centro
                c.border = borda_branca
        else:
            # Zebra nas demais colunas (B:E)
            fill_linha = fill_branco if row % 2 == 0 else fill_cinza
            for col in range(2, max_col + 1):
                c = ws.cell(row=row, column=col)
                c.fill = fill_linha
                c.font = fonte_preta_negrito
                c.alignment = alinhamento_centro
                c.border = borda_padrao

        # Colunas de percentual com formato de porcentagem (1 casa)
        for col_perc in (4, 5):
            c = ws.cell(row=row, column=col_perc)
            if isinstance(c.value, (int, float)):
                c.value = c.value / 100
            c.number_format = "0.0%"

    # Ajuste de largura das colunas com minimo para ja abrir "pronto para leitura"
    larguras_minimas = {
        1: 28,  # ESPECIALISTA
        2: 38,  # TOTAL DE CIRURGIAS REALIZADAS
        3: 16,  # Nº BENEF "SIM"
        4: 20,  # PROPORCIONALIDADE
        5: 20,  # REPRESENTATIVIDADE
    }

    for col in range(1, max_col + 1):
        letra = get_column_letter(col)
        maior = 0
        for row in range(1, max_row + 1):
            valor = ws.cell(row=row, column=col).value
            tamanho = len(str(valor)) if valor is not None else 0
            if tamanho > maior:
                maior = tamanho
        largura_auto = maior + 4
        ws.column_dimensions[letra].width = max(largura_auto, larguras_minimas.get(col, 12))

    ws.freeze_panes = "A2"
def formatar_tabela(
    arquivo_entrada: str,
    arquivo_saida: str,
    abas: list[str] | None = None,
) -> None:
    wb = load_workbook(arquivo_entrada)

    abas_para_formatar = abas if abas else wb.sheetnames
    for nome_aba in abas_para_formatar:
        if nome_aba in wb.sheetnames:
            _formatar_aba(wb[nome_aba])

    wb.save(arquivo_saida)


if __name__ == "__main__":
    arquivo_origem = "indicadores_calculados.xlsx"
    arquivo_destino = "indicadores_calculados_formatado.xlsx"
    formatar_tabela(arquivo_origem, arquivo_destino)
    print(f"Arquivo formatado salvo em: {arquivo_destino}")
