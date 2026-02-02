from openpyxl import load_workbook

# ============================================================
# CONFIGURAÇÕES
# ============================================================
arquivo_entrada = "BASE DEZEMBRO INTERNACAO - CHECKPOINT APAGAR DUPLICADOS.xlsx"
arquivo_saida = "BASE_DEZEMBRO_INTERNACAO_SEM_DUPLICADOS.xlsx"
aba_alvo = "BASE"
coluna_status = "STATUS DUPLICADO"

print("🔹 Iniciando processo de exclusão de duplicados...")
print(f"📂 Arquivo de entrada: {arquivo_entrada}")

# ============================================================
# 1) ABRINDO O ARQUIVO
# ============================================================
wb = load_workbook(arquivo_entrada)
print("✅ Arquivo carregado com sucesso")

# ============================================================
# 2) ACESSANDO A ABA BASE
# ============================================================
if aba_alvo not in wb.sheetnames:
    raise ValueError(f"❌ Aba '{aba_alvo}' não encontrada no arquivo")

ws = wb[aba_alvo]
print(f"📄 Aba '{aba_alvo}' acessada")

# ============================================================
# 3) LOCALIZANDO A COLUNA 'STATUS DUPLICADO'
# ============================================================
header = {}
for col in range(1, ws.max_column + 1):
    valor = ws.cell(row=1, column=col).value
    if valor:
        header[valor.strip()] = col

if coluna_status not in header:
    raise ValueError(f"❌ Coluna '{coluna_status}' não encontrada na aba BASE")

col_status_idx = header[coluna_status]
print(f"🔎 Coluna '{coluna_status}' encontrada na posição {col_status_idx}")

# ============================================================
# 4) IDENTIFICANDO LINHAS PARA EXCLUSÃO
# ============================================================
linhas_para_excluir = []

for row in range(2, ws.max_row + 1):
    valor = ws.cell(row=row, column=col_status_idx).value
    if valor == "APAGAR DUPLICADO":
        linhas_para_excluir.append(row)

print(f"🧮 Total de linhas marcadas para exclusão: {len(linhas_para_excluir)}")

# ============================================================
# 5) EXCLUINDO LINHAS (DE BAIXO PARA CIMA)
# ============================================================
if linhas_para_excluir:
    for row in reversed(linhas_para_excluir):
        ws.delete_rows(row)

    print("🗑️ Linhas excluídas com sucesso")
else:
    print("⚠️ Nenhuma linha encontrada para exclusão")

# ============================================================
# 6) SALVANDO NOVO ARQUIVO
# ============================================================
wb.save(arquivo_saida)
print(f"💾 Novo arquivo gerado: {arquivo_saida}")

print("🏁 Processo finalizado com sucesso!")
